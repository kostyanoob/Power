import os
import pandas as pd
def printLogHeader(logFile, val_en = True):
    '''
    Prints a nice header to a log file
    '''
    valStr = 'Val ' if val_en else 'Test'
    logFile.write("|-------||-----------|------------|---------------|----------------|\n")
    logFile.write("| Epoch || " + valStr + " Loss | Train Loss | " + valStr + " Acc. (%) | Train Acc. (%) |\n")
    logFile.write("|-------||-----------|------------|---------------|----------------|\n")

def printLogEntry(logFile, logLine, vl, tl, va, ta):
    '''
    Prints one log entry containing validation loss (vl)
    training loss (tl), validation accuracy (va) and validateion
    accuracy (va)
    '''
    qualifier = "| {0:5d} ||".format(int(logLine)) if type(logLine)==int else "|" + str(logLine) + "||"
    logFile.write(qualifier + '   {0:7.4f} |   {1:8.4f} | {2:13f} | {3:14f} |\n'.format(float(vl),
                                                                                        float(tl),
                                                                                        float(va),
                                                                                        float(ta)))
def printPernodeMSEHeader(logFile, element_titles_list):
    '''
    Prints a nice header to a log file
    '''
    col_width=12
    logFile.write("|-------||"+ "".join(["------------|" for _ in element_titles_list]) + "\n")
    logFile.write("| Epoch ||" + "".join([" "*((col_width-len(ite))//2)+ite+" "*((col_width-len(ite))//2)+"|" for ite in element_titles_list]) + "\n")
    logFile.write("|-------||"+ "".join(["------------|" for _ in element_titles_list]) + "\n")

def printPernodeEntry(logFile, logLine, numbers_list_or_vector):
    '''
    Prints one log entry containing validation loss (vl)
    training loss (tl), validation accuracy (va) and validateion
    accuracy (va)
    '''
    # import pdb
    # pdb.set_trace()
    col_width = 12
    if type(numbers_list_or_vector) != type([]):
        element_titles_list = numbers_list_or_vector.tolist()
    else:
        element_titles_list = numbers_list_or_vector

    qualifier = "| {0:5d} ||".format(int(logLine)) if type(logLine) == int else "|" + str(logLine) + "||"
    logFile.write(qualifier + "".join(["  {0:8.6f}  |".format(ite) for ite in element_titles_list])+ "\n")


def is_log_test_MSE_per_example_complete(dir, log_filename):
    """
    Note, this function parses the test_MSE per example log file,
    which is normally terminated by the following 6 lines:
    Test MSE of Re normalized-average-across-buses 3.629e-02
    Test MSE of Im normalized-average-across-buses 5.467e-03
    Test MSE of All normalized-average-across-buses 2.088e-02
    Test MSE of Magnitude denormalized-average-across-buses 3.672e-02
    Test MSE of Angle denormalized-average-across-buses 1.702e+01
    Test MSE of All denormalized-average-across-buses 8.528e+00


    :param dir: a directory to the log files
    :param log_filename: the name of the log file including the extension ".txt"
    :return: True <--> log file is complete
    """
    fullpath = os.path.join(dir, log_filename)
    if not os.path.exists(fullpath):
        return False
    with open(fullpath, "r") as file:
        lines = file.readlines()
        if len(lines) < 6:
            return False
        else:
            all_lines_appear = "Test MSE of All normalized-average-across-buses" in lines[-4] \
                               and "Test MSE of Im normalized-average-across-buses" in lines[-5] \
                               and "Test MSE of Re normalized-average-across-buses" in lines[-6] \
                               and "Test MSE of All denormalized-average-across-buses" in lines[-1] \
                               and "Test MSE of Angle denormalized-average-across-buses" in lines[-2] \
                               and "Test MSE of Magnitude denormalized-average-across-buses" in lines[-3]

            return all_lines_appear


def is_log_complete(dir, log_filename):
    """
    Opens dir/log_filename and check that the log file contains the
    words "Test accuracy" at its last line. If these words are indeed there,
    then it means that the log file was completed.

    :param dir: a directory to the log files
    :param log_filename: the name of the log file including the extension ".txt"
    :return: True <--> log file is complete
    """
    fullpath = os.path.join(dir, log_filename)
    if not os.path.exists(fullpath):
        return False
    with open(fullpath, "r") as file:
        lines = file.readlines()
        if len(lines) < 1:
            return False
        else:
            return "Test MSE" in lines[-1]


def write_summary(summary_writer, summary_str, summary_writer_epoch_counter):
    summary_writer.add_summary(summary_str, summary_writer_epoch_counter)
    summary_writer.flush()
    summary_writer_epoch_counter += 1
    return summary_writer, summary_writer_epoch_counter

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()